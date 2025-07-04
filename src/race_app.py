import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import datetime
import asyncio
import threading
import csv
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import numpy as np

from src.race_parser import RaceParser
from src.locale_manager import LocaleManager

class RaceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.locale = LocaleManager()
        self.title(self.locale.tr('app_title'))
        self.geometry("1000x700")
        self.resizable(True, True)
        self.participants_data = []
        self.debug_mode = tk.BooleanVar(value=False)
        self.all_pages_mode = tk.BooleanVar(value=False)
        self.current_date = datetime.date.today()
        self.create_menu()
        self.create_widgets()
    
    def create_menu(self):
        self.menubar = tk.Menu(self)
        
        self.file_menu = tk.Menu(self.menubar, tearoff=0)
        self.file_menu.add_command(
            label=self.locale.tr('exit_menu'), 
            command=self.destroy
        )
        self.menubar.add_cascade(
            label=self.locale.tr('file_menu'), 
            menu=self.file_menu
        )
        
        self.language_menu = tk.Menu(self.menubar, tearoff=0)
        self.language_menu.add_command(
            label=self.locale.tr('russian'), 
            command=lambda: self.change_language('ru')
        )
        self.language_menu.add_command(
            label=self.locale.tr('english'), 
            command=lambda: self.change_language('en')
        )
        self.language_menu.add_command(
            label=self.locale.tr('chinese'), 
            command=lambda: self.change_language('cn')
        )
        self.language_menu.add_command(
            label=self.locale.tr('japanese'), 
            command=lambda: self.change_language('jp')
        )
        self.menubar.add_cascade(
            label=self.locale.tr('language_menu'), 
            menu=self.language_menu
        )
        
        self.config(menu=self.menubar)
    
    def change_language(self, language):
        self.locale.set_language(language)
        self.title(self.locale.tr('app_title'))
        self.update_ui_texts()
        self.config(menu=None)
        self.create_menu()
    
    def update_ui_texts(self):
        self.input_frame.config(text=self.locale.tr('input_params'))
        self.race_type_label.config(text=self.locale.tr('race_type'))
        self.year_label.config(text=self.locale.tr('year'))
        self.month_label.config(text=self.locale.tr('month'))
        self.day_label.config(text=self.locale.tr('day'))
        self.week_label.config(text=self.locale.tr('week'))
        self.page_label.config(text=self.locale.tr('page'))
        self.export_type_label.config(text=self.locale.tr('export_type'))
        self.load_button.config(text=self.locale.tr('load_data'))
        self.export_button.config(text=self.locale.tr('export_data'))
        self.debug_check.config(text=self.locale.tr('debug_mode'))
        self.all_pages_check.config(text=self.locale.tr('all_pages'))
        self.search_label.config(text=self.locale.tr('search_participant'))
        self.search_button.config(text=self.locale.tr('search'))
        
        self.tree.heading("rank", text=self.locale.tr('rank'))
        self.tree.heading("name", text=self.locale.tr('participant'))
        self.tree.heading("distance", text=self.locale.tr('distance'))
        self.tree.heading("kills", text=self.locale.tr('kills'))
        
        self.status_var.set(self.locale.tr('status_ready'))
        
        self.notebook.tab(0, text=self.locale.tr('data_tab'))
        self.notebook.tab(1, text=self.locale.tr('analysis_tab'))
        
        self.analysis_frame.config(text=self.locale.tr('graph_controls'))
        self.graph_type_label.config(text=self.locale.tr('graph_type'))
        self.top_count_label.config(text=self.locale.tr('top_count'))
        self.plot_button.config(text=self.locale.tr('plot_graph'))
        self.save_button.config(text=self.locale.tr('save_graph'))
        self.analysis_status.set(self.locale.tr('analysis_ready'))
        
        graph_types = [
            self.locale.tr('distance_distribution'),
            self.locale.tr('kills_distribution'),
            self.locale.tr('top_distance'),
            self.locale.tr('top_kills'),
            self.locale.tr('distance_vs_kills')
        ]
        self.graph_type['values'] = graph_types
        self.graph_type.current(0)
    
    def create_widgets(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        main_frame = ttk.Frame(self.notebook, padding=10)
        self.notebook.add(main_frame, text=self.locale.tr('data_tab'))
        
        self.input_frame = ttk.LabelFrame(main_frame, text=self.locale.tr('input_params'))
        self.input_frame.pack(fill=tk.X, pady=5)

        self.race_type_label = ttk.Label(self.input_frame, text=self.locale.tr('race_type'))
        self.race_type_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        self.race_type = ttk.Combobox(self.input_frame, values=["daily", "weekly"], state="readonly")
        self.race_type.current(0)
        self.race_type.grid(row=0, column=1, padx=5, pady=5)

        self.year_label = ttk.Label(self.input_frame, text=self.locale.tr('year'))
        self.year_label.grid(row=0, column=2, padx=5, pady=5, sticky=tk.W)
        self.year = ttk.Entry(self.input_frame, width=8)
        self.year.grid(row=0, column=3, padx=5, pady=5)
        self.year.insert(0, str(self.current_date.year))
        
        self.month_label = ttk.Label(self.input_frame, text=self.locale.tr('month'))
        self.month_label.grid(row=0, column=4, padx=5, pady=5, sticky=tk.W)
        self.month = ttk.Entry(self.input_frame, width=5)
        self.month.grid(row=0, column=5, padx=5, pady=5)
        self.month.insert(0, f"{self.current_date.month:02d}")
        
        self.day_label = ttk.Label(self.input_frame, text=self.locale.tr('day'))
        self.day_label.grid(row=0, column=6, padx=5, pady=5, sticky=tk.W)
        self.day = ttk.Entry(self.input_frame, width=5)
        self.day.grid(row=0, column=7, padx=5, pady=5)
        self.day.insert(0, f"{self.current_date.day:02d}")
        
        self.week_label = ttk.Label(self.input_frame, text=self.locale.tr('week'))
        self.week_label.grid(row=1, column=4, padx=5, pady=5, sticky=tk.W)
        self.week = ttk.Entry(self.input_frame, width=5)
        self.week.grid(row=1, column=5, padx=5, pady=5)
        week_num = self.current_date.isocalendar()[1]
        self.week.insert(0, str(week_num))
        
        self.page_label = ttk.Label(self.input_frame, text=self.locale.tr('page'))
        self.page_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        self.page = ttk.Entry(self.input_frame, width=5)
        self.page.insert(0, "1")
        self.page.grid(row=1, column=1, padx=5, pady=5)

        self.load_button = ttk.Button(
            self.input_frame, 
            text=self.locale.tr('load_data'), 
            command=self.load_data
        )
        self.load_button.grid(row=1, column=8, padx=5, pady=5)

        self.debug_check = ttk.Checkbutton(
            self.input_frame, 
            text=self.locale.tr('debug_mode'), 
            variable=self.debug_mode
        )
        self.debug_check.grid(row=1, column=9, padx=10, pady=5)

        self.all_pages_check = ttk.Checkbutton(
            self.input_frame, 
            text=self.locale.tr('all_pages'), 
            variable=self.all_pages_mode
        )
        self.all_pages_check.grid(row=1, column=10, padx=10, pady=5)

        self.export_type_label = ttk.Label(self.input_frame, text=self.locale.tr('export_type'))
        self.export_type_label.grid(row=0, column=8, padx=5, pady=5, sticky=tk.W)
        self.export_type = ttk.Combobox(self.input_frame, values=["xlsx", "csv", "txt"], state="readonly", width=5)
        self.export_type.current(0)
        self.export_type.grid(row=0, column=9, padx=5, pady=5)

        self.export_button = ttk.Button(
            self.input_frame, 
            text=self.locale.tr('export_data'), 
            command=self.export_data
        )
        self.export_button.grid(row=0, column=10, padx=5, pady=5)

        self.status_var = tk.StringVar(value=self.locale.tr('status_ready'))
        status_label = ttk.Label(self.input_frame, textvariable=self.status_var)
        status_label.grid(row=2, column=0, columnspan=14, sticky=tk.W, padx=5, pady=5)
        
        search_frame = ttk.Frame(main_frame)
        search_frame.pack(fill=tk.X, pady=5)
        
        self.search_label = ttk.Label(search_frame, text=self.locale.tr('search_participant'))
        self.search_label.pack(side=tk.LEFT, padx=5)
        self.search_entry = ttk.Entry(search_frame, width=30)
        self.search_entry.pack(side=tk.LEFT, padx=5)
        self.search_button = ttk.Button(search_frame, text=self.locale.tr('search'), command=self.search_participant)
        self.search_button.pack(side=tk.LEFT, padx=5)

        columns = ("rank", "name", "distance", "kills")
        self.tree = ttk.Treeview(main_frame, columns=columns, show="headings")
        
        self.tree.heading("rank", text=self.locale.tr('rank'))
        self.tree.heading("name", text=self.locale.tr('participant'))
        self.tree.heading("distance", text=self.locale.tr('distance'))
        self.tree.heading("kills", text=self.locale.tr('kills'))
        
        self.tree.column("rank", width=50, anchor=tk.CENTER)
        self.tree.column("name", width=200)
        self.tree.column("distance", width=80, anchor=tk.CENTER)
        self.tree.column("kills", width=80, anchor=tk.CENTER)
        
        scrollbar = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.race_type.bind("<<ComboboxSelected>>", self.update_input_fields)
        self.update_input_fields()
        
        analysis_tab = ttk.Frame(self.notebook)
        self.notebook.add(analysis_tab, text=self.locale.tr('analysis_tab'))
        self.create_analysis_tab(analysis_tab)
    
    def create_analysis_tab(self, parent):
        self.analysis_frame = ttk.LabelFrame(parent, text=self.locale.tr('graph_controls'))
        self.analysis_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.graph_type_label = ttk.Label(self.analysis_frame, text=self.locale.tr('graph_type'))
        self.graph_type_label.grid(row=0, column=0, padx=5, pady=5)
        self.graph_type = ttk.Combobox(self.analysis_frame, state="readonly", width=20)
        self.graph_type.grid(row=0, column=1, padx=5, pady=5)
        
        self.top_count_label = ttk.Label(self.analysis_frame, text=self.locale.tr('top_count'))
        self.top_count_label.grid(row=0, column=2, padx=5, pady=5)
        self.top_count = ttk.Spinbox(self.analysis_frame, from_=5, to=100, width=5)
        self.top_count.set(20)
        self.top_count.grid(row=0, column=3, padx=5, pady=5)
        
        self.plot_button = ttk.Button(
            self.analysis_frame, 
            text=self.locale.tr('plot_graph'),
            command=self.plot_graph
        )
        self.plot_button.grid(row=0, column=4, padx=10, pady=5)
        
        self.analysis_status = tk.StringVar(value=self.locale.tr('analysis_ready'))
        ttk.Label(self.analysis_frame, textvariable=self.analysis_status).grid(
            row=0, column=5, padx=10, pady=5, sticky=tk.W
        )
        
        self.graph_frame = ttk.Frame(parent)
        self.graph_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        self.figure = plt.figure(figsize=(8, 5), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.figure, master=self.graph_frame)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        save_frame = ttk.Frame(parent)
        save_frame.pack(fill=tk.X, padx=10, pady=5)
        
        self.save_button = ttk.Button(
            save_frame,
            text=self.locale.tr('save_graph'),
            command=self.save_graph
        )
        self.save_button.pack(side=tk.RIGHT, padx=5, pady=5)
    
    def update_input_fields(self, event=None):
        race_type = self.race_type.get()
        if race_type == "daily":
            self.week.config(state=tk.DISABLED)
            self.month.config(state=tk.NORMAL)
            self.day.config(state=tk.NORMAL)
        else:
            self.week.config(state=tk.NORMAL)
            self.month.config(state=tk.DISABLED)
            self.day.config(state=tk.DISABLED)
    
    def load_data(self):
        threading.Thread(
            target=self.run_async_task, 
            args=(self.async_load_data(),),
            daemon=True
        ).start()
    
    def run_async_task(self, coro):
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        try:
            loop.run_until_complete(coro)
        except Exception as e:
            self.after(0, lambda: messagebox.showerror(
                self.locale.tr('error_title'), 
                f"Error: {str(e)}"
            ))
        finally:
            loop.close()
    
    async def async_load_data(self):
        race_type = self.race_type.get()
        year = self.year.get().strip()
        page = self.page.get().strip()
        debug = self.debug_mode.get()
        all_pages = self.all_pages_mode.get()
        
        if debug:
            print("\n" + "="*50)
            print(f"Loading data: {datetime.datetime.now()}")
            print(f"All pages mode: {all_pages}")
        
        if not year:
            self.after(0, lambda: messagebox.showwarning(
                self.locale.tr('error_title'), 
                self.locale.tr('year_required')
            ))
            return
        
        if race_type == "daily":
            month = self.month.get().strip()
            day = self.day.get().strip()
            if not month or not day:
                self.after(0, lambda: messagebox.showwarning(
                    self.locale.tr('error_title'), 
                    self.locale.tr('month_day_required')
                ))
                return
            identifier = (month, day)
        else:
            week = self.week.get().strip()
            if not week:
                self.after(0, lambda: messagebox.showwarning(
                    self.locale.tr('error_title'), 
                    self.locale.tr('week_required')
                ))
                return
            identifier = week

        self.participants_data = []
        
        if all_pages:
            self.update_status(self.locale.tr('loading_all'))
            page = 1
            total_participants = 0
            
            while True:
                if debug:
                    print(f"Loading page {page}...")
                
                self.update_status(self.locale.tr('loading_page', page=page))
                
                try:
                    participants = await RaceParser.parse_race(
                        race_type, year, identifier, page, debug
                    )
                except Exception as e:
                    self.update_status(str(e))
                    break
                
                if participants is None:
                    break
                
                if not participants:
                    if debug:
                        print(f"Page {page} is empty, finishing")
                    break
                
                self.participants_data.extend(participants)
                total_participants += len(participants)
                
                if debug:
                    print(f"Page {page}: loaded {len(participants)} participants")
                
                page += 1
            
            self.update_status(self.locale.tr(
                'loaded_participants', 
                count=total_participants, 
                pages=page-1
            ))
        else:
            if not page:
                self.after(0, lambda: messagebox.showwarning(
                    self.locale.tr('error_title'), 
                    self.locale.tr('page_required')
                ))
                return
            
            self.update_status(self.locale.tr('loading_page', page=page))
            
            try:
                participants = await RaceParser.parse_race(race_type, year, identifier, page, debug)
                if participants is not None:
                    self.participants_data = participants
                    self.update_status(self.locale.tr('loaded_participants', 
                        count=len(participants), 
                        pages=1
                    ))
            except Exception as e:
                self.update_status(str(e))
        
        if self.participants_data:
            self.update_display()
            if debug:
                print(f"Total participants loaded: {len(self.participants_data)}")
        else:
            self.update_status(self.locale.tr('no_participants'))
    
    def update_status(self, message):
        self.after(0, lambda: self.status_var.set(message))
    
    def update_display(self):
        self.after(0, self.display_current_data)
    
    def display_current_data(self):
        self.display_data(self.participants_data)
    
    def export_data(self):
        if not self.participants_data:
            messagebox.showwarning(
                self.locale.tr('no_data'), 
                self.locale.tr('no_data_message')
            )
            return
        
        file_type = self.export_type.get()

        file_path = filedialog.asksaveasfilename(
            defaultextension=f".{file_type}",
            filetypes=[(f"{file_type.upper()} files", f"*.{file_type}")],
            title=self.locale.tr('export_data')
        )
        
        if not file_path:
            return
        
        try:
            if file_type == "xlsx":
                self.export_to_xlsx(file_path)
            elif file_type == "csv":
                self.export_to_csv(file_path)
            elif file_type == "txt":
                self.export_to_txt(file_path)
                
            messagebox.showinfo(
                self.locale.tr('export_success'),
                self.locale.tr('export_success_message', file_path=file_path)
            )
        except Exception as e:
            messagebox.showerror(
                self.locale.tr('export_error'),
                f"{self.locale.tr('export_error')}: {str(e)}"
            )
    
    def export_to_xlsx(self, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.locale.tr('participant_results')

        headers = [
            self.locale.tr('rank'),
            self.locale.tr('participant'),
            self.locale.tr('distance'),
            self.locale.tr('kills')
        ]
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        for participant in self.participants_data:
            ws.append([
                participant['rank'],
                participant['name'],
                participant['distance'],
                participant['kills']
            ])
        
        wb.save(file_path)
    
    def export_to_csv(self, file_path):
        with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            
            writer.writerow([
                self.locale.tr('rank'),
                self.locale.tr('participant'),
                self.locale.tr('distance'),
                self.locale.tr('kills')
            ])
            
            for participant in self.participants_data:
                writer.writerow([
                    participant['rank'],
                    participant['name'],
                    participant['distance'],
                    participant['kills']
                ])

    def export_to_txt(self, file_path):
        with open(file_path, 'w', newline='', encoding='utf-8') as txtfile:            
            txtfile.write(
                self.locale.tr('rank') + '\t' +
                self.locale.tr('participant') + '\t' +
                self.locale.tr('distance') + '\t' +
                self.locale.tr('kills') + '\n'
            )
            
            for participant in self.participants_data:
                txtfile.write(
                    participant['rank'] + '\t' +
                    participant['name'] + '\t' +
                    participant['distance'] + '\t' +
                    participant['kills'] + '\n'
                )
    
    def display_data(self, data):
        self.tree.delete(*self.tree.get_children())
        for item in data:
            self.tree.insert("", tk.END, values=(
                item['rank'],
                item['name'],
                item['distance'],
                item['kills']
            ))
    
    def search_participant(self):
        query = self.search_entry.get().lower().strip()
        if not query:
            self.display_data(self.participants_data)
            return
        
        filtered = [p for p in self.participants_data if query in p['name'].lower()]
        self.display_data(filtered)
    
    def plot_graph(self):
        if not self.participants_data:
            self.analysis_status.set(self.locale.tr('no_data_for_analysis'))
            return
            
        try:
            self.figure.clear()
            graph_type = self.graph_type.get()
            top_count = int(self.top_count.get())
            
            if graph_type == self.locale.tr('distance_distribution'):
                self.plot_distance_distribution()
            elif graph_type == self.locale.tr('kills_distribution'):
                self.plot_kills_distribution()
            elif graph_type == self.locale.tr('top_distance'):
                self.plot_top_distance(top_count)
            elif graph_type == self.locale.tr('top_kills'):
                self.plot_top_kills(top_count)
            elif graph_type == self.locale.tr('distance_vs_kills'):
                self.plot_distance_vs_kills(top_count)
                
            self.canvas.draw()
            self.analysis_status.set(self.locale.tr('graph_plotted'))
        except Exception as e:
            self.analysis_status.set(f"{self.locale.tr('graph_error')}: {str(e)}")
    
    def plot_distance_distribution(self):
        ax = self.figure.add_subplot(111)
        distances = [self.parse_number(p['distance']) for p in self.participants_data]
        distances = [d for d in distances if d > 0]
        if not distances:
            raise ValueError(self.locale.tr('no_valid_distance'))
        ax.hist(distances, bins=100, color='skyblue', edgecolor='black')
        ax.set_title(self.locale.tr('distance_distribution'))
        ax.set_xlabel(self.locale.tr('distance'))
        ax.set_ylabel(self.locale.tr('participant_count'))
        ax.grid(True, linestyle='--', alpha=0.7)
    
    def plot_kills_distribution(self):
        ax = self.figure.add_subplot(111)
        kills = [self.parse_number(p['kills']) for p in self.participants_data]
        kills = [k for k in kills if k > 0]
        if not kills:
            raise ValueError(self.locale.tr('no_valid_kills'))
        ax.hist(kills, bins=30, color='salmon', edgecolor='black')
        ax.set_title(self.locale.tr('kills_distribution'))
        ax.set_xlabel(self.locale.tr('kills'))
        ax.set_ylabel(self.locale.tr('participant_count'))
        ax.grid(True, linestyle='--', alpha=0.7)
    
    def plot_top_distance(self, top_count):
        sorted_data = sorted(
            self.participants_data,
            key=lambda x: self.parse_number(x['distance']),
            reverse=True
        )[:top_count]
        names = [p['name'][:20] + ('...' if len(p['name']) > 20 else '') for p in sorted_data]
        distances = [self.parse_number(p['distance']) for p in sorted_data]
        ax = self.figure.add_subplot(111)
        y_pos = np.arange(len(names))
        ax.barh(y_pos, distances, color='lightgreen')
        ax.set_yticks(y_pos)
        ax.set_yticklabels(names)
        ax.invert_yaxis()
        ax.set_title(self.locale.tr('top_distance', count=top_count))
        ax.set_xlabel(self.locale.tr('distance'))
        ax.grid(True, axis='x', linestyle='--', alpha=0.7)
    
    def plot_top_kills(self, top_count):
        sorted_data = sorted(
            self.participants_data,
            key=lambda x: self.parse_number(x['kills']),
            reverse=True
        )[:top_count]
        names = [p['name'][:20] + ('...' if len(p['name']) > 20 else '') for p in sorted_data]
        kills = [self.parse_number(p['kills']) for p in sorted_data]
        ax = self.figure.add_subplot(111)
        y_pos = np.arange(len(names))
        ax.barh(y_pos, kills, color='gold')
        ax.set_yticks(y_pos)
        ax.set_yticklabels(names)
        ax.invert_yaxis()
        ax.set_title(self.locale.tr('top_kills', count=top_count))
        ax.set_xlabel(self.locale.tr('kills'))
        ax.grid(True, axis='x', linestyle='--', alpha=0.7)
    
    def plot_distance_vs_kills(self, top_count):
        sorted_data = sorted(
            self.participants_data,
            key=lambda x: self.parse_number(x['distance']),
            reverse=True
        )[:top_count]
        names = [p['name'][:15] + ('...' if len(p['name']) > 15 else '') for p in sorted_data]
        distances = [self.parse_number(p['distance']) for p in sorted_data]
        kills = [self.parse_number(p['kills']) for p in sorted_data]
        ax = self.figure.add_subplot(111)
        x_pos = np.arange(len(names))
        width = 0.35
        ax.bar(x_pos - width/2, distances, width, label=self.locale.tr('distance'), color='skyblue')
        ax.bar(x_pos + width/2, kills, width, label=self.locale.tr('kills'), color='salmon')
        ax.set_xticks(x_pos)
        ax.set_xticklabels(names, rotation=45, ha='right')
        ax.set_title(self.locale.tr('distance_vs_kills', count=top_count))
        ax.legend()
        ax.grid(True, linestyle='--', alpha=0.7)
    
    def parse_number(self, value):
        if isinstance(value, (int, float)):
            return float(value)
        try:
            cleaned = ''.join(ch for ch in str(value) if ch.isdigit() or ch in ['.', '-'])
            return float(cleaned) if cleaned else 0.0
        except ValueError:
            return 0.0
    
    def save_graph(self):
        if not self.participants_data:
            messagebox.showwarning(
                self.locale.tr('no_data'),
                self.locale.tr('no_data_for_analysis')
            )
            return
            
        file_path = filedialog.asksaveasfilename(
            defaultextension=".png",
            filetypes=[
                ("PNG files", "*.png"),
                ("JPEG files", "*.jpg"),
                ("PDF files", "*.pdf"),
                ("All files", "*.*")
            ],
            title=self.locale.tr('save_graph')
        )
        
        if not file_path:
            return
            
        try:
            self.figure.savefig(file_path, bbox_inches='tight', dpi=100)
            messagebox.showinfo(
                self.locale.tr('success'),
                self.locale.tr('graph_saved', path=file_path)
            )
        except Exception as e:
            messagebox.showerror(
                self.locale.tr('error'),
                f"{self.locale.tr('save_error')}: {str(e)}"
            )
